# -*- coding: utf-8 -*-
"""Excel 导出模块。

使用 openpyxl 生成包含三个子表的工作簿：
1. “原始数据”：上传文件中的原始数据（未清洗）；
2. “处理后的数据”：清洗/插值后的数据 + 插值标记；
3. “图片”：渲染好的曲线图图片 + 原生可编辑 Excel 图表
   （双击可编辑、修改数据后图表自动更新）。
"""

from __future__ import annotations

import io
import math

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import Paragraph, RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties, Font as XLFont, ParagraphProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .cleaning import CleaningStats, FLAG_INTERPOLATED, FLAG_NAMES, FLAG_ORIGINAL, FLAG_UNFIXED
from .config import (
    COLOR_CASING,
    COLOR_GAS,
    COLOR_OIL,
    FONT_NAME,
    LINEWIDTH_CASING,
    LINEWIDTH_GAS,
    LINEWIDTH_OIL,
    get_logger,
)

logger = get_logger()

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, color="808080")
# 修改数据标色：插值修复 → 浅黄色；无法修复 → 浅红色
FILL_INTERPOLATED = PatternFill("solid", fgColor="FFF2CC")
FILL_UNFIXED = PatternFill("solid", fgColor="F4CCCC")

# “处理后的数据”Sheet 中：A=日期，B=油压，C=套压，D=瞬时气量（清洗后）
COL_DATE = 1
COL_OIL = 2
COL_CASING = 3
COL_GAS = 4

# 图表字体与字号（方正大黑简体，7-9 号；图例大 1 个字号）
CHART_FONT = FONT_NAME
AXIS_TITLE_FONT_SIZE = 900    # 9pt（DrawingML 单位为百分之一磅）
TICK_LABEL_FONT_SIZE = 800    # 8pt
LEGEND_FONT_SIZE = 1000       # 10pt（比坐标轴标签大 1 号）
AXIS_LINE_WIDTH_EMU = 6350    # 0.5 磅 = 6350 EMU


def _char_props(
    size: int,
    font_name: str = CHART_FONT,
    bold: bool | None = None,
) -> CharacterProperties:
    """构造字符属性：指定字体、字号，可显式控制是否加粗。"""
    return CharacterProperties(
        sz=size,
        b=bold,
        latin=XLFont(typeface=font_name),
        ea=XLFont(typeface=font_name),
        cs=XLFont(typeface=font_name),
        lang="zh-CN",
    )


def _style_axis(axis) -> None:
    """设置坐标轴：外部黑色 0.5 磅刻度线、黑色 0.5 磅轴线、刻度文字字体字号。"""
    axis.majorTickMark = "out"
    axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="000000", w=AXIS_LINE_WIDTH_EMU))
    axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=_char_props(TICK_LABEL_FONT_SIZE)))]
    )
    if axis.title is not None and getattr(axis.title, "tx", None) is not None:
        rich = getattr(axis.title.tx, "rich", None)
        if rich is not None:
            for p in rich.p:
                # 纵轴标题不设置加粗（b=0），与普通字体保持一致
                p.pPr = ParagraphProperties(defRPr=_char_props(AXIS_TITLE_FONT_SIZE, bold=False))
                p.endParaRPr = _char_props(AXIS_TITLE_FONT_SIZE, bold=False)
                for r in p.r:
                    r.rPr = _char_props(AXIS_TITLE_FONT_SIZE, bold=False)


def _style_legend(legend) -> None:
    """设置图例字体：比坐标轴标签大 1 个字号（10 号）。"""
    legend.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=_char_props(LEGEND_FONT_SIZE)))]
    )


def _vertical_title(axis) -> None:
    """纵轴标题文字方向：竖排（从右到左）。"""
    if axis.title is not None and getattr(axis.title, "tx", None) is not None:
        rich = getattr(axis.title.tx, "rich", None)
        if rich is not None and rich.bodyPr is not None:
            rich.bodyPr.vert = "eaVert"
            rich.bodyPr.rtlCol = True
            rich.bodyPr.rot = 0


def _write_raw_sheet(ws, df_raw: pd.DataFrame) -> None:
    """写入“原始数据”Sheet：文件中的原始数据，不做任何清洗。"""
    columns = ["日期", "油压", "套压", "瞬时气量"]
    for j, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, row in df_raw.iterrows():
        excel_row = i + 2
        ws.cell(row=excel_row, column=1, value=row["日期"]).number_format = "yyyy-mm-dd hh:mm:ss"
        for j, col in enumerate(columns[1:], start=2):
            value = row[col]
            if col == "瞬时气量" and pd.notna(value):
                value = round(float(value), 4)
            cell = ws.cell(row=excel_row, column=j, value=value)
            if col == "瞬时气量":
                cell.number_format = "0.0000"
    ws.freeze_panes = "A2"
    for j in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18
    note = ws.cell(row=len(df_raw) + 3, column=1, value="本表为上传文件的原始数据（未清洗）。")
    note.font = NOTE_FONT


def _write_clean_sheet(ws, df_clean: pd.DataFrame, stats: CleaningStats) -> None:
    """写入“处理后的数据”Sheet：清洗/插值后的数据 + 插值标记。

    被修改（插值修复）的数据以浅黄色标出，无法修复的空值以浅红色标出。
    """
    columns = [
        "日期",
        "油压", "套压", "瞬时气量",
        "油压_插值标记", "套压_插值标记", "瞬时气量_插值标记",
    ]
    for j, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    data_cols = ["油压", "套压", "瞬时气量"]
    for i, row in df_clean.iterrows():
        excel_row = i + 2
        ws.cell(row=excel_row, column=1, value=row["日期"]).number_format = "yyyy-mm-dd hh:mm:ss"
        for j, col in enumerate(data_cols, start=2):
            flag = int(row[f"{col}_插值标记"]) if pd.notna(row[f"{col}_插值标记"]) else FLAG_ORIGINAL
            cell = ws.cell(row=excel_row, column=j, value=row[col])
            if col == "瞬时气量":
                cell.number_format = "0.0000"
            # 插值修复的数据标浅黄色
            if flag == FLAG_INTERPOLATED:
                cell.fill = FILL_INTERPOLATED
            # 插值标记列：插值→浅黄，无法修复→浅红
            flag_cell = ws.cell(row=excel_row, column=j + 3, value=FLAG_NAMES.get(flag, str(flag)))
            if flag == FLAG_INTERPOLATED:
                flag_cell.fill = FILL_INTERPOLATED
            elif flag == FLAG_UNFIXED:
                flag_cell.fill = FILL_UNFIXED
    ws.freeze_panes = "A2"
    for j in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18

    note_row = len(df_clean) + 3
    notes = [
        "插值标记说明：0=原始有效，1=插值修复，2=无法修复（保留空值）",
        f"数据量：{stats.total_rows} 行；时间范围：{stats.time_min} ~ {stats.time_max}；"
        f"采样频率：{stats.frequency}；瞬时气量为 0 的行数：{stats.zero_gas_rows}",
    ]
    for k, note in enumerate(notes):
        cell = ws.cell(row=note_row + k, column=1, value=note)
        cell.font = NOTE_FONT


def _excel_serial(ts: pd.Timestamp) -> float:
    """把时间转为 Excel 序列号（1899-12-30 为基准）。"""
    return (ts - pd.Timestamp("1899-12-30")).total_seconds() / 86400.0


def _integer_step(value_range: float) -> int:
    """为纵轴选择整数刻度步长，使刻度节点控制在 4~6 个。

    取“不小于 最大值/5”的最小 1/2/5×10^n 整数步长，保证刻度不超过 6 个。
    """
    if value_range <= 0:
        return 1
    ideal = value_range / 5.0
    exp = 10 ** math.floor(math.log10(max(ideal, 1e-9)))
    for m in (1, 2, 5, 10):
        if m * exp >= ideal:
            return int(m * exp)
    return int(10 * exp)


def _y_axis_range(series: pd.Series) -> tuple[float, int, int]:
    """根据数据计算纵轴：最低 0、最高（+10% 取整）、整数刻度步长。"""
    valid = pd.to_numeric(series, errors="coerce").dropna()
    if valid.empty:
        return 0.0, 1, 1
    top = max(1.0, float(valid.max()) * 1.1)
    # 向上取整到整数
    top = math.ceil(top)
    step = _integer_step(top)
    top = math.ceil(top / step) * step
    return 0.0, int(top), step


def _build_native_chart(
    data_ws,
    df_clean: pd.DataFrame,
    stats: CleaningStats,
):
    """构建原生可编辑的 Excel 图表（左轴压力、右轴瞬时气量）。"""
    n_rows = len(df_clean)
    xvalues = Reference(data_ws, min_col=COL_DATE, min_row=2, max_row=n_rows + 1)
    left_top = _y_axis_range(pd.concat([df_clean["油压"], df_clean["套压"]]))
    right_top = _y_axis_range(df_clean["瞬时气量"])

    def add_series(
        chart,
        col: int,
        title: str,
        color_hex: str,
        width_pt: float,
    ) -> None:
        yvalues = Reference(data_ws, min_col=col, min_row=2, max_row=n_rows + 1)
        ser = Series(yvalues, xvalues, title=title)
        ser.marker = Marker(symbol="none")
        ser.graphicalProperties = GraphicalProperties(
            ln=LineProperties(solidFill=color_hex.lstrip("#"), w=int(width_pt * 12700))
        )
        chart.series.append(ser)

    # 三条曲线均为散点折线，保持真实日期间距；左轴压力、右轴瞬时气量
    chart_left = ScatterChart()
    chart_left.scatterStyle = "line"
    chart_left.legend.position = "t"  # 图例在顶部，横向单行
    chart_left.x_axis.axPos = "b"
    chart_left.x_axis.majorGridlines = None
    chart_left.y_axis.majorGridlines = None
    chart_left.y_axis.title = "压力（MPa）"
    add_series(chart_left, COL_OIL, "油压", COLOR_OIL, LINEWIDTH_OIL)
    add_series(chart_left, COL_CASING, "套压", COLOR_CASING, LINEWIDTH_CASING)
    _style_axis(chart_left.x_axis)
    _style_axis(chart_left.y_axis)
    _vertical_title(chart_left.y_axis)
    chart_left.y_axis.scaling.min = left_top[0]
    chart_left.y_axis.scaling.max = left_top[1]
    chart_left.y_axis.majorUnit = left_top[2]
    chart_left.y_axis.number_format = "0"

    chart_right = ScatterChart()
    chart_right.scatterStyle = "line"
    chart_right.x_axis.axId = 10
    chart_right.x_axis.crossAx = 200
    chart_right.y_axis.axId = 200
    chart_right.y_axis.crossAx = 10
    chart_right.y_axis.axPos = "r"
    chart_right.y_axis.crosses = "max"
    chart_right.y_axis.majorGridlines = None
    chart_right.y_axis.title = "瞬时气量（万方/天）"
    add_series(chart_right, COL_GAS, "瞬时气量", COLOR_GAS, LINEWIDTH_GAS)
    _style_axis(chart_right.x_axis)
    _style_axis(chart_right.y_axis)
    _vertical_title(chart_right.y_axis)
    chart_right.y_axis.scaling.min = right_top[0]
    chart_right.y_axis.scaling.max = right_top[1]
    chart_right.y_axis.majorUnit = right_top[2]
    chart_right.y_axis.number_format = "0"

    chart = chart_left
    chart += chart_right

    # X 轴：日期格式 + 等时间间隔刻度（6 个，间隔 = 跨度/5）
    tmin, tmax = stats.time_min, stats.time_max
    span_days = (tmax - tmin).total_seconds() / 86400.0
    chart.x_axis.number_format = "yyyy/m/d"
    chart.x_axis.majorUnit = span_days / 5.0
    chart.x_axis.scaling.min = _excel_serial(tmin)
    chart.x_axis.scaling.max = _excel_serial(tmax)

    _style_legend(chart.legend)
    # 图表边框：默认黑色
    chart.graphical_properties = GraphicalProperties(
        ln=LineProperties(solidFill="000000", w=9525)
    )

    # 图表尺寸：12 × 6 英寸 ≈ 30.5 × 15.3 厘米
    chart.width = 30.5
    chart.height = 15.3
    return chart


def _write_chart_sheet(
    ws,
    template_name: str,
    well_name: str,
    stats: CleaningStats,
    chart: ScatterChart,
) -> None:
    """写入“图片”Sheet：信息行 + 原生可编辑图表（不嵌入静态图片）。"""
    ws["A1"] = (
        f"模板：{template_name}；井号：{well_name}；数据量：{stats.total_rows} 条；"
        f"时间范围：{stats.time_min} ~ {stats.time_max}；采样频率：{stats.frequency}"
    )
    hint_row = 2
    if stats.warnings:
        ws["A2"] = "处理说明：" + "；".join(stats.warnings)
        ws["A2"].font = NOTE_FONT
        ws["A2"].alignment = Alignment(wrap_text=True)
        hint_row = 3
    ws.cell(
        row=hint_row, column=1,
        value="下方为原生 Excel 图表，双击即可编辑；修改「处理后的数据」子表中的数值，图表会自动更新。",
    ).font = NOTE_FONT

    # 原生可编辑图表（引用“处理后的数据”子表）
    ws.add_chart(chart, "A5")


def export_excel(
    df_raw: pd.DataFrame,
    df_clean: pd.DataFrame,
    stats: CleaningStats,
    well_name: str,
    template_name: str = "单井生产曲线模板",
) -> bytes:
    """生成包含三个子表的 Excel 文件，返回字节内容。"""
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "原始数据"
    _write_raw_sheet(ws_raw, df_raw)

    ws_clean = wb.create_sheet("处理后的数据")
    _write_clean_sheet(ws_clean, df_clean, stats)

    ws_chart = wb.create_sheet("图片")
    chart = _build_native_chart(ws_clean, df_clean, stats)
    _write_chart_sheet(ws_chart, template_name, well_name, stats, chart)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("Excel 导出完成：三个子表（原始数据/处理后的数据/图片），原生图表已嵌入。")
    return buf.getvalue()
