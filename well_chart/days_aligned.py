# -*- coding: utf-8 -*-
"""第二模块：天数叠合（时间拉齐）曲线模板。

处理流程：
1. 自动识别列名（日期/油压/套压/瞬时气量，支持变体）；
2. 数据清洗：负值修复、滑动窗口固定值检测、气量仪表故障修复；
3. 时间拉齐：按生产天数组织数据（同一天多条记录取平均），天数从 1 递增；
4. 输出 Excel：时间拉齐数据 / 清洗日志 / 曲线图（原生可编辑折线图）。

本模块使用面向对象实现，颜色、字号等常量集中在 config.py。
"""

from __future__ import annotations

import io
import math
from dataclasses import dataclass, field

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import NumericAxis
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .cleaning import interpolate_series
from .config import (
    MOD2_CHART_HEIGHT,
    MOD2_CHART_WIDTH,
    MOD2_COLOR_CASING,
    MOD2_COLOR_GAS,
    MOD2_COLOR_OIL,
    MOD2_FIXED_POINTS,
    MOD2_FIXED_STD,
    MOD2_FIXED_WINDOW,
    MOD2_GAS_MIN_POINTS,
    MOD2_GAS_STD,
    MOD2_LINEWIDTH,
    MOD2_MAX_ROWS,
    MOD2_NEG_CASING_FACTOR,
    MOD2_NEG_GAS,
    MOD2_NEG_OIL_FACTOR,
    MOD2_SAMPLE_ROWS,
    MOD2_X_TICKS,
    get_logger,
)
from .excel_export import _style_axis, _style_legend
from .loader import _parse_dates

logger = get_logger()

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)


# ---------------------------------------------------------------------------
# 清洗日志
# ---------------------------------------------------------------------------
@dataclass
class CleaningLog:
    """记录所有数据清洗操作，供“清洗日志”Sheet 展示。"""

    records: list = field(default_factory=list)

    def add(self, category: str, message: str) -> None:
        """追加一条清洗日志（同时输出到控制台日志）。"""
        self.records.append(
            {
                "序号": len(self.records) + 1,
                "时间": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                "类别": category,
                "说明": message,
            }
        )
        logger.info("[%s] %s", category, message)

    def to_dataframe(self) -> pd.DataFrame:
        """把日志转为 DataFrame；无记录时返回带表头的空表。"""
        if not self.records:
            return pd.DataFrame(columns=["序号", "时间", "类别", "说明"])
        return pd.DataFrame(self.records)


# ---------------------------------------------------------------------------
# 列名识别
# ---------------------------------------------------------------------------
class ColumnDetector:
    """自动识别日期/油压/套压/瞬时气量列（支持常见列名变体）。"""

    @staticmethod
    def _norm(value) -> str:
        """标准化单元格文本：转小写并去掉首尾空白。"""
        return "" if value is None else str(value).strip().lower()

    @staticmethod
    def detect(raw: pd.DataFrame, header_row: int) -> dict | None:
        """在表头行及其上方两行中查找四列位置。

        识别规则：
        - 日期列：包含 日期/时间/date/time/timestamp/datetime；
        - 油压列：同时包含“油”和“压”；
        - 套压列：同时包含“套”和“压”；
        - 气量列：包含 气/量/gas。
        全部识别成功返回 {date, oil, casing, gas}，否则返回 None。
        """
        result: dict[str, int] = {}
        for i in range(max(0, header_row - 2), header_row + 1):
            for j in range(raw.shape[1]):
                text = ColumnDetector._norm(raw.iat[i, j])
                if not text:
                    continue
                kind = None
                if "套" in text and "压" in text:
                    kind = "casing"
                elif "油" in text and "压" in text:
                    kind = "oil"
                elif any(k in text for k in ("气", "量", "gas")):
                    kind = "gas"
                elif any(k in text for k in ("日期", "时间", "date", "time", "timestamp", "datetime")):
                    kind = "date"
                if kind is not None and kind not in result:
                    result[kind] = j
        if all(k in result for k in ("date", "oil", "casing", "gas")):
            logger.info("列名识别成功：%s", result)
            return result
        return None

    @staticmethod
    def detect_aligned(raw: pd.DataFrame, header_row: int) -> dict | None:
        """识别“时间拉齐”格式的数据源。

        识别规则（关键词）：
        - 天数：包含“天”；
        - 平均套压(兆帕)：包含“套”；
        - 平均油压(兆帕)：包含“油”；
        - 平均日产气(万方)：包含“气”。
        全部识别成功返回 {day, casing, oil, gas}，否则返回 None。
        """
        result: dict[str, int] = {}
        for i in range(max(0, header_row - 2), header_row + 1):
            for j in range(raw.shape[1]):
                text = ColumnDetector._norm(raw.iat[i, j])
                if not text:
                    continue
                kind = None
                if "天" in text:
                    kind = "day"
                elif "套" in text:
                    kind = "casing"
                elif "油" in text:
                    kind = "oil"
                elif "气" in text:
                    kind = "gas"
                if kind is not None and kind not in result:
                    result[kind] = j
        if all(k in result for k in ("day", "casing", "oil", "gas")):
            logger.info("时间拉齐格式识别成功：%s", result)
            return result
        return None


def build_std_df(raw: pd.DataFrame, header_row: int, cols: dict) -> tuple[pd.DataFrame, CleaningLog]:
    """按识别出的列构建标准化 DataFrame（日期/油压/套压/瞬时气量）。

    日期无法解析的行会被跳过并记录日志。
    """
    log = CleaningLog()
    data = raw.iloc[header_row + 1:].copy()
    df = pd.DataFrame(
        {
            "日期": _parse_dates(data.iloc[:, cols["date"]]),
            "油压": pd.to_numeric(data.iloc[:, cols["oil"]], errors="coerce"),
            "套压": pd.to_numeric(data.iloc[:, cols["casing"]], errors="coerce"),
            "瞬时气量": pd.to_numeric(data.iloc[:, cols["gas"]], errors="coerce"),
        }
    )
    bad_dates = df["日期"].isna()
    if bad_dates.any():
        log.add("日期解析", f"{int(bad_dates.sum())} 行日期无法解析，已跳过")
        df = df[~bad_dates]
    df = df.sort_values("日期").reset_index(drop=True)
    if df.empty:
        raise ValueError("解析后没有有效数据（日期列均无法解析）。")
    return df, log


def build_aligned_df(raw: pd.DataFrame, header_row: int, cols: dict) -> tuple[pd.DataFrame, CleaningLog]:
    """读取已为“时间拉齐”格式的数据源（天数/平均套压/平均油压/平均日产气）。

    天数无法解析的行会被跳过并记录日志，数值统一保留 4 位小数。
    """
    log = CleaningLog()
    data = raw.iloc[header_row + 1:].copy()
    df = pd.DataFrame(
        {
            "天数": pd.to_numeric(data.iloc[:, cols["day"]], errors="coerce"),
            "平均套压": pd.to_numeric(data.iloc[:, cols["casing"]], errors="coerce"),
            "平均油压": pd.to_numeric(data.iloc[:, cols["oil"]], errors="coerce"),
            "平均日产气": pd.to_numeric(data.iloc[:, cols["gas"]], errors="coerce"),
        }
    )
    bad_days = df["天数"].isna()
    if bad_days.any():
        log.add("数据识别", f"{int(bad_days.sum())} 行天数无法解析，已跳过")
        df = df[~bad_days]
    df = df.sort_values("天数").reset_index(drop=True)
    df[["平均套压", "平均油压", "平均日产气"]] = df[["平均套压", "平均油压", "平均日产气"]].round(4)
    if df.empty:
        raise ValueError("解析后没有有效数据（天数列均无法解析）。")
    log.add("数据识别", "文件已为“时间拉齐”格式（天数/平均套压/平均油压/平均日产气），直接作为数据源使用")
    return df, log


# ---------------------------------------------------------------------------
# 数据清洗
# ---------------------------------------------------------------------------
def detect_fixed_runs(
    series: pd.Series,
    points: int = MOD2_FIXED_POINTS,
    std_thr: float = MOD2_FIXED_STD,
) -> list:
    """滑动窗口法检测固定值段。

    以连续 points 个点为窗口计算标准差，若标准差 < std_thr 则把该窗口内
    的点标记为无效；把相邻的无效点合并成段返回 [(起点, 终点), ...]。
    """
    s = pd.to_numeric(series, errors="coerce")
    rolling_std = s.rolling(points, min_periods=points).std()
    hits = (rolling_std < std_thr).fillna(False).to_numpy()
    invalid = np.zeros(len(s), dtype=bool)
    for i in np.where(hits)[0]:
        invalid[max(0, i - points + 1): i + 1] = True

    runs: list = []
    in_run = False
    start = 0
    for i, flag in enumerate(invalid):
        if flag and not in_run:
            in_run, start = True, i
        elif not flag and in_run:
            in_run = False
            runs.append((start, i - 1))
    if in_run:
        runs.append((start, len(invalid) - 1))
    return runs


class DataCleaner:
    """数据清洗：负值修复、压力固定值插值、气量仪表故障修复。"""

    def __init__(self, log: CleaningLog):
        self.log = log

    def clean(self, df: pd.DataFrame) -> pd.DataFrame:
        """执行完整清洗流程（按日期排序后原地修改）。"""
        df = df.copy().sort_values("日期").reset_index(drop=True)
        self._fix_negatives(df)
        for col in ("油压", "套压"):
            self._fix_constant_pressure(df, col)
        self._fix_gas_meter(df)
        return df

    def _fix_negatives(self, df: pd.DataFrame) -> None:
        """负值处理：气量→0；油压→前值×0.9；套压→前值×0.95。"""
        gas_neg = df["瞬时气量"] < 0
        if gas_neg.any():
            df.loc[gas_neg, "瞬时气量"] = MOD2_NEG_GAS
            self.log.add("负值处理", f"瞬时气量 {int(gas_neg.sum())} 个负值已替换为 {MOD2_NEG_GAS}")
        for col, factor in (("油压", MOD2_NEG_OIL_FACTOR), ("套压", MOD2_NEG_CASING_FACTOR)):
            mask = df[col] < 0
            if mask.any():
                prev = df[col].shift(1)
                has_prev = mask & prev.notna()
                df.loc[has_prev, col] = prev[has_prev] * factor
                df.loc[mask & prev.isna(), col] = 0.0
                self.log.add(
                    "负值处理",
                    f"{col} {int(mask.sum())} 个负值已按“前值×{factor}”修复，无前值时置 0",
                )

    def _fix_constant_pressure(self, df: pd.DataFrame, col: str) -> None:
        """压力固定值检测：滑动窗口法，固定段置空后按时间线性插值。"""
        runs = detect_fixed_runs(df[col])
        if not runs:
            return
        total_points = sum(j - i + 1 for i, j in runs)
        for i, j in runs:
            df.loc[i:j, col] = np.nan
        flags = np.zeros(len(df), dtype=np.int8)
        cleaned, _, unfixed = interpolate_series(df[col], df["日期"], flags)
        df[col] = cleaned
        self.log.add(
            "固定值修复",
            f"{col} 检测到 {len(runs)} 段固定值（窗口 {MOD2_FIXED_WINDOW}，"
            f"连续 {MOD2_FIXED_POINTS} 点标准差<{MOD2_FIXED_STD}），共 {total_points} 点，"
            f"已线性插值；无法修复 {int(unfixed)} 点",
        )

    def _fix_gas_meter(self, df: pd.DataFrame) -> None:
        """气量仪表故障检测：固定正值（段内标准差<0.05 且超过 10 个点）用均值替换。"""
        s = pd.to_numeric(df["瞬时气量"], errors="coerce")
        # 只对正值做固定值检测：非正值置 NaN 以切断滑动窗口
        work = s.mask(s <= 0)
        rolling_std = work.rolling(MOD2_FIXED_POINTS, min_periods=MOD2_FIXED_POINTS).std()
        hits = (rolling_std < MOD2_GAS_STD).fillna(False).to_numpy()
        invalid = np.zeros(len(s), dtype=bool)
        for i in np.where(hits)[0]:
            invalid[max(0, i - MOD2_FIXED_POINTS + 1): i + 1] = True

        runs: list = []
        in_run = False
        start = 0
        for i, flag in enumerate(invalid):
            if flag and not in_run:
                in_run, start = True, i
            elif not flag and in_run:
                in_run = False
                runs.append((start, i - 1))
        if in_run:
            runs.append((start, len(invalid) - 1))
        # 只处理点数超过阈值的固定段
        runs = [r for r in runs if r[1] - r[0] + 1 > MOD2_GAS_MIN_POINTS]
        if not runs:
            return
        mask = np.zeros(len(s), dtype=bool)
        for i, j in runs:
            mask[i:j + 1] = True
        valid = s.to_numpy(dtype=float)[~mask]
        mean_val = float(valid.mean()) if len(valid) and np.isfinite(valid).any() else 0.0
        df.loc[mask, "瞬时气量"] = round(mean_val, 4)
        self.log.add(
            "气量仪表修复",
            f"检测到 {len(runs)} 段固定正值（标准差<{MOD2_GAS_STD} 且超过 "
            f"{MOD2_GAS_MIN_POINTS} 点），已用均值 {mean_val:.4f} 替换",
        )


# ---------------------------------------------------------------------------
# 时间拉齐
# ---------------------------------------------------------------------------
class TimeAligner:
    """按生产天数重新组织数据：同一天多条记录取平均，天数从 1 递增。"""

    @staticmethod
    def align(df: pd.DataFrame) -> pd.DataFrame:
        """返回列：天数、平均套压、平均油压、平均日产气（均保留 4 位小数）。"""
        work = df.copy()
        work["_day"] = work["日期"].dt.normalize()
        grouped = (
            work.groupby("_day", sort=True)[["套压", "油压", "瞬时气量"]]
            .mean()
        )
        return pd.DataFrame(
            {
                "天数": np.arange(1, len(grouped) + 1),
                "平均套压": grouped["套压"].round(4).to_numpy(),
                "平均油压": grouped["油压"].round(4).to_numpy(),
                "平均日产气": grouped["瞬时气量"].round(4).to_numpy(),
            }
        )


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
class DaysAlignedProcessor:
    """第二模块主流程：采样 → 清洗 → 时间拉齐。"""

    def __init__(
        self,
        df: pd.DataFrame,
        well_name: str = "未知井号",
        log: CleaningLog | None = None,
        pre_aligned: bool = False,
    ):
        self.df = df
        self.well_name = well_name
        self.log = log if log is not None else CleaningLog()
        self.pre_aligned = pre_aligned

    def run(self) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
        """执行完整流程，返回 (时间拉齐数据, 清洗日志, 统计信息)。"""
        if self.pre_aligned:
            aligned = self.df[["天数", "平均套压", "平均油压", "平均日产气"]].round(4).reset_index(drop=True)
            self.log.add("数据识别", "数据源已为时间拉齐格式，跳过清洗与拉齐，直接出图")
            stats = {
                "days": len(aligned),
                "raw_rows": len(aligned),
                "log_count": len(self.log.records),
                "well_name": self.well_name,
            }
            logger.info("时间拉齐数据直接使用：共 %d 天", stats["days"])
            return aligned, self.log.to_dataframe(), stats
        df = self._sample_if_needed(self.df)
        df = DataCleaner(self.log).clean(df)
        aligned = TimeAligner.align(df)
        stats = {
            "days": len(aligned),
            "raw_rows": len(df),
            "log_count": len(self.log.records),
            "well_name": self.well_name,
        }
        logger.info("时间拉齐完成：共 %d 个生产天数，清洗操作 %d 条", stats["days"], stats["log_count"])
        return aligned, self.log.to_dataframe(), stats

    def _sample_if_needed(self, df: pd.DataFrame) -> pd.DataFrame:
        """数据量超过上限时均匀采样，保证性能。"""
        if len(df) <= MOD2_MAX_ROWS:
            return df.reset_index(drop=True)
        indices = np.linspace(0, len(df) - 1, MOD2_SAMPLE_ROWS).round().astype(int)
        out = df.iloc[np.unique(indices)].sort_values("日期").reset_index(drop=True)
        self.log.add(
            "数据采样",
            f"数据量 {len(df)} 超过 {MOD2_MAX_ROWS}，已均匀采样至 {len(out)} 行",
        )
        return out


# ---------------------------------------------------------------------------
# 横坐标刻度
# ---------------------------------------------------------------------------
def tick_positions(n: int, num_ticks: int = MOD2_X_TICKS) -> list:
    """计算 6 个等距离天数刻度（包含起点 1 和终点 n）。"""
    if n <= 1:
        return [1]
    step = max(1, round((n - 1) / (num_ticks - 1)))
    ticks = [1 + i * step for i in range(num_ticks)]
    ticks[-1] = n  # 保证最后一个刻度是总天数
    out: list = []
    for t in ticks:
        if not out or t != out[-1]:
            out.append(t)
    return out


def preview_figure(aligned: pd.DataFrame):
    """生成网页预览图（仅界面展示用；下载的 Excel 使用原生图表）。"""
    import matplotlib.pyplot as plt

    from .plotting import _setup_chinese_fonts

    _setup_chinese_fonts()
    fig, ax = plt.subplots(figsize=(10, 4.8), dpi=100, facecolor="white")
    ax2 = ax.twinx()
    x = aligned["天数"]
    ax.plot(x, aligned["平均套压"], color=MOD2_COLOR_CASING, linewidth=1.2, label="平均套压")
    ax.plot(x, aligned["平均油压"], color=MOD2_COLOR_OIL, linewidth=1.2, label="平均油压")
    ax2.plot(x, aligned["平均日产气"], color=MOD2_COLOR_GAS, linewidth=1.2, label="平均日产气")

    ax.set_xlabel("天数")
    ax.set_ylabel("压力（MPa）")
    ax2.set_ylabel("日产气（万方）")
    ax.set_xticks(tick_positions(len(aligned)))
    ax.set_xlim(1, max(1, len(aligned)))
    ax.grid(False)
    ax2.grid(False)

    lines = ax.get_lines() + ax2.get_lines()
    labels = [line.get_label() for line in lines]
    ax.legend(lines, labels, loc="upper center", ncol=3, frameon=False)
    ax.spines["top"].set_visible(False)
    ax2.spines["top"].set_visible(False)
    fig.tight_layout()
    return fig


# ---------------------------------------------------------------------------
# Excel 原生折线图
# ---------------------------------------------------------------------------
def build_native_chart(data_ws, aligned: pd.DataFrame) -> LineChart:
    """构建原生可编辑 Excel 折线图（左轴压力、右轴日产气、X 轴为天数）。"""
    n = len(aligned)
    x_ref = Reference(data_ws, min_col=1, min_row=2, max_row=n + 1)
    width_emu = int(MOD2_LINEWIDTH * 12700)  # 2.25 磅

    def add_series(chart: LineChart, col: int, title: str, color_hex: str) -> None:
        """向折线图添加一条带样式的序列（图例不带单位，只显示名称）。"""
        ref = Reference(data_ws, min_col=col, min_row=2, max_row=n + 1)
        ser = Series(ref, title=title)
        ser.marker = Marker(symbol="none")
        ser.graphicalProperties = GraphicalProperties(
            ln=LineProperties(solidFill=color_hex.lstrip("#"), w=width_emu)
        )
        chart.series.append(ser)

    def vertical_title(axis) -> None:
        """纵轴标题竖排：从上往下读（与模板一致）。"""
        if axis.title is not None and getattr(axis.title, "tx", None) is not None:
            rich = getattr(axis.title.tx, "rich", None)
            if rich is not None and rich.bodyPr is not None:
                rich.bodyPr.vert = "eaVert"

    # 左轴图表：平均套压 + 平均油压（压力）
    chart_left = LineChart()
    chart_left.x_axis = NumericAxis(axId=10, crossAx=20)
    chart_left.x_axis.axPos = "b"
    chart_left.y_axis.axId = 20
    chart_left.y_axis.crossAx = 10
    chart_left.y_axis.majorGridlines = None
    chart_left.y_axis.title = "压力（兆帕）"
    add_series(chart_left, 2, "平均套压", MOD2_COLOR_CASING)
    add_series(chart_left, 3, "平均油压", MOD2_COLOR_OIL)
    chart_left.set_categories(x_ref)
    _style_axis(chart_left.x_axis)
    _style_axis(chart_left.y_axis)
    vertical_title(chart_left.y_axis)

    # 右轴图表：平均日产气（产量，独立右纵轴）
    chart_right = LineChart()
    chart_right.x_axis.axId = 10
    chart_right.x_axis.crossAx = 200
    chart_right.y_axis.axId = 200
    chart_right.y_axis.crossAx = 10
    chart_right.y_axis.axPos = "r"
    chart_right.y_axis.crosses = "max"
    chart_right.y_axis.majorGridlines = None
    chart_right.y_axis.title = "日产气（万方）"
    add_series(chart_right, 4, "平均日产气", MOD2_COLOR_GAS)
    chart_right.set_categories(x_ref)
    _style_axis(chart_right.x_axis)
    _style_axis(chart_right.y_axis)
    vertical_title(chart_right.y_axis)

    chart = chart_left
    chart += chart_right

    # X 轴：天数，固定 6 个等距离刻度（1 → 总天数）
    chart.x_axis.number_format = "0"
    chart.x_axis.majorUnit = max(1.0, (n - 1) / (MOD2_X_TICKS - 1))
    chart.x_axis.scaling.min = 1
    chart.x_axis.scaling.max = max(1, n)

    # 纵轴：最小 0，最大按数据 × 1.1 / × 1.15 向上取整
    pressure_max = max(
        float(aligned["平均套压"].max()), float(aligned["平均油压"].max())
    )
    gas_max = float(aligned["平均日产气"].max())
    chart_left.y_axis.scaling.min = 0
    chart_left.y_axis.scaling.max = max(1, math.ceil(pressure_max * 1.1))
    chart_left.y_axis.number_format = "0.00"
    chart_right.y_axis.scaling.min = 0
    chart_right.y_axis.scaling.max = max(1, math.ceil(gas_max * 1.15))
    chart_right.y_axis.number_format = "0.00"

    # 图例：顶部居中、单行、无边框、字体大 2 号
    chart.legend.position = "t"
    chart.legend.spPr = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
    _style_legend(chart.legend)

    # 图表：无标题、无边框、尺寸 15 × 8 厘米
    chart.width = MOD2_CHART_WIDTH
    chart.height = MOD2_CHART_HEIGHT
    return chart


# ---------------------------------------------------------------------------
# Excel 导出（三个子表）
# ---------------------------------------------------------------------------
class DaysExcelExporter:
    """导出“时间拉齐数据 / 清洗日志 / 曲线图”三个子表的 Excel。"""

    DATA_HEADERS = ["天数", "平均套压(兆帕)", "平均油压(兆帕)", "平均日产气(万方)"]
    LOG_HEADERS = ["序号", "时间", "类别", "说明"]

    @staticmethod
    def _write_header(ws, headers: list) -> None:
        """写表头：加粗、居中、浅蓝底。"""
        for j, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=j, value=text)
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.fill = HEADER_FILL

    @staticmethod
    def _auto_width(ws, headers: list, rows: pd.DataFrame) -> None:
        """按内容长度设置列宽（自适应）。"""
        for j, header in enumerate(headers, start=1):
            width = len(str(header))
            for value in rows.iloc[:, j - 1]:
                if value is not None:
                    width = max(width, len(str(value)))
            ws.column_dimensions[get_column_letter(j)].width = min(width + 2, 40)

    def export(
        self,
        aligned: pd.DataFrame,
        log_df: pd.DataFrame,
        well_name: str,
    ) -> bytes:
        """生成 Excel 字节内容。"""
        wb = Workbook()

        # Sheet1：时间拉齐数据
        ws1 = wb.active
        ws1.title = "时间拉齐数据"
        self._write_header(ws1, self.DATA_HEADERS)
        for i, row in aligned.iterrows():
            excel_row = i + 2
            ws1.cell(row=excel_row, column=1, value=int(row["天数"]))
            for j, col in enumerate(("平均套压", "平均油压", "平均日产气"), start=2):
                cell = ws1.cell(row=excel_row, column=j, value=float(row[col]))
                cell.number_format = "0.0000"
        self._auto_width(ws1, self.DATA_HEADERS, aligned)
        ws1.freeze_panes = "A2"

        # Sheet2：清洗日志
        ws2 = wb.create_sheet("清洗日志")
        self._write_header(ws2, self.LOG_HEADERS)
        for i, row in log_df.iterrows():
            excel_row = i + 2
            for j, col in enumerate(self.LOG_HEADERS, start=1):
                ws2.cell(row=excel_row, column=j, value=row[col])
        if log_df.empty:
            ws2.cell(row=2, column=4, value="无清洗操作")
        self._auto_width(ws2, self.LOG_HEADERS, log_df)
        ws2.freeze_panes = "A2"

        # Sheet3：曲线图（原生可编辑折线图）
        ws3 = wb.create_sheet("曲线图")
        chart = build_native_chart(ws1, aligned)
        ws3.add_chart(chart, "A1")
        ws3["A17"] = f"井号：{well_name}；生产天数：{len(aligned)} 天"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        logger.info("第二模块 Excel 导出完成：三个子表，原生折线图已嵌入")
        return buf.getvalue()
